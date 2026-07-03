"""
Export de la liste des badges scannés vers CSV ou Excel.
"""

import csv
from datetime import datetime


def exporter_csv(badges: list, chemin: str) -> bool:
    """
    Exporte une liste de badges vers un fichier CSV.
    badges : liste de tuples (numero, badge_id, heure)
    """
    try:
        with open(chemin, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(["#", "Numéro de badge", "Utilisateur", "Heure", "Statut"])
            for ligne in badges:
                writer.writerow(ligne)
        return True
    except OSError:
        return False


def exporter_excel(badges: list, chemin: str) -> bool:
    """
    Exporte une liste de badges vers un fichier Excel (.xlsx).
    Nécessite openpyxl. Retourne False si la librairie est absente
    ou en cas d'erreur d'écriture.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Badges scannés"

        entetes = ["#", "Numéro de badge", "Utilisateur", "Heure", "Statut"]
        ws.append(entetes)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F6BFF", end_color="4F6BFF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for ligne in badges:
            ws.append(ligne)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16

        wb.save(chemin)
        return True
    except OSError:
        return False


def nom_fichier_horodate(prefixe: str, extension: str) -> str:
    horodatage = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"{prefixe}_{horodatage}.{extension}"
